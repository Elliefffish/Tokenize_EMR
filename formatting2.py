import pandas as pd
import threading
import openpyxl
import aes_org

Gender = {'F':0, 'M':1}
County = {}
Identity = {}
Major = {}
School = {}
entry_day = ''
class Worker(threading.Thread):
    def __init__(self, entry_day, infile, outfile):
        threading.Thread.__init__(self)
        self.infile = infile
        self.outfile = outfile
        self.InTitle = {}
        self.OutTitle = {}
        self.outcols = []
        self.entry_day = entry_day
    def run(self):
        if(self.infile.split('.')[-1] == 'xlsx'):
            self.read_xlsx()
            self.write_xlsx()
        elif(self.infile.split('.')[-1] == 'csv'):
            self.read_csv()
            self.write_csv()
    def read_xlsx(self):
        self.wb_in = openpyxl.load_workbook(self.infile)
        wb_out = openpyxl.load_workbook(self.outfile)
        self.ws_in = self.wb_in.active
        ws_out = wb_out.active
        # Convert the header of infile into a dict "InTitle"
        self.InTitle = [i.value for i in list(self.ws_in.rows)[0]]
        # Convert the header of outfile into a dict "OutTitle"
        for i in range(0, len(list(ws_out.rows)[1])):
            self.OutTitle[list(ws_out.rows)[1][i].value] = i
            # Init a empty list "outcols" for each ouput row
            self.outcols.append(None)
#        print(self.OutTitle)
    def write_xlsx(self):
        wb = openpyxl.load_workbook(self.outfile)
        ws = wb.active
        rf1 = pd.DataFrame(pd.read_excel(self.infile)).to_csv("input.csv",  index = None, header=True, encoding='utf-8')
        with open('input.csv', 'r') as readFile:
            line = readFile.readline()
            while True:
                fields = readFile.readline()[:-1].split(',')
                if len(fields)==1: break
                for i in range(0, len(fields)):
                    self.convert(self.InTitle[i], fields[i])
                ws.append(self.outcols)
        wb.save('static/學生基本資料上傳.xlsx')
    def read_csv(self):
        # read header of infile and outfile.
        with open(infile, 'r') as readFile, open(outfile, 'r', newline='') as writeFile:
            # Convert the header of infile into a dict "InTitle"
            self.InTitle = readFile.readline()[:-1].split(',')
            # Convert the header of outfile into a dict "OutTitle"
            line = writeFile.readline()
            line = writeFile.readline()[:-1].split(',')
            for col in range(0, len(line)):
                if line[col]!='':
                    self.OutTitle[line[col]] = col
            # Init a empty list "outcols" for each ouput row
            self.outcols = ['' for i in range(0, len(line))]
    def write_csv(self):
        with open(self.infile, 'r') as readFile, open(self.outfile, 'a', newline='') as writeFile:
            line = readFile.readline()
            while True:
                fields = readFile.readline()[:-1].split(',')
                if len(fields)==1: break
                for i in range(0, len(fields)):
                    self.convert(self.InTitle[i], fields[i])
                print(*self.outcols, sep=',', file=writeFile)
        #out = pd.DataFrame(pd.read_csv("output.csv")).to_excel(self.infile+"_new.xlsx",  index = None, header=True)
        out = pd.DataFrame(pd.read_csv("output.csv")).to_excel("static/output.xlsx",  index = None, header=True)

    def birth(self, data):
        try:
            data = str(data)
            self.outcols[self.OutTitle['出生年月']] ='{}/{}/{}'.format(int(data[:-4])+1911, data[-4:-2], data[-2:])
        except:
            self.outcols[self.OutTitle['出生年月']] = -998

    def entry_date(self, data):
        try:
            self.outcols[self.OutTitle['學生開學日期']] = self.entry_day.replace('-', '/')
            data = str(data)
            if (len(data) < 6 and len(data) > 3):
                entry_day = '{}/{}'.format(int(data[:-2])+1911, data[-2:])
            elif (len(data) >= 6):
                entry_day = '{}/{}/{}'.format(int(data[:-4])+1911, data[-4:-2], data[-2:])
        except:
            self.outcols[self.OutTitle['學生開學日期']] = -998
    def gender(self, data):
        try:
           self.outcols[self.OutTitle['性別']] = Gender[data]
        except KeyError:
            self.outcols[self.OutTitle['性別']] = -998
    def county(self, data):
        try:
            data = data.lstrip('0123456789')[:3]
            if data[0]=='台':
                data='臺'+data[1:3]
            if data in County:
                self.outcols[self.OutTitle['戶籍縣市']] = County[data]
            else: self.outcols[self.OutTitle['戶籍縣市']] = County[data[:2]+'縣']
        except:
            self.outcols[self.OutTitle['戶籍縣市']] = -998
    def nationality(self, data):
        id_num = self.outcols[self.OutTitle['學生身份別']]
        if data == 'ROC':
            self.outcols[self.OutTitle['是否為本國籍學生']] = 1
            self.outcols[self.OutTitle['學生國籍']] = '臺灣'
            if id_num != 2:
                self.outcols[self.OutTitle['學生身份別']] = 1
        else:
            self.outcols[self.OutTitle['是否為本國籍學生']] = 0
            self.outcols[self.OutTitle['學生國籍']] = data
    def identity(self, data):
        try:
            data = data.split('-')
            for item in data:
                if item == '原住民子女(H)':
                    self.outcols[self.OutTitle['學生身份別']] = 2
                    return
                elif item == '外國學生':
                    self.outcols[self.OutTitle['學生身份別']] = 4
                    return
                elif item == '僑生':
                    self.outcols[self.OutTitle['學生身份別']] = 5
                    return
            self.outcols[self.OutTitle['學生身份別']] = -998
        except:
            self.outcols[self.OutTitle['學生身份別']] = -998
    def status(self, data):
        data = data.rstrip(')')[-1]
        self.outcols[self.OutTitle['在校狀況']] = int(data)

    def department(self, data):
        try:
            if data < 11 and data > 3:
                self.outcols[self.OutTitle['學生學制']] = data
            else:
                Dict = {'學士':4, '碩士':10, '博士':9}
                self.outcols[self.OutTitle['學生學制']] = Dict[str(data)[:2]]
        except:
            self.outcols[self.OutTitle['學生學制']] = -998
    def en_id(self, data):
        self.outcols[self.OutTitle['學校加密流水號']] = aes_org.aes_encrypt(str(data), 'ncnu_studentid')
    def major(self, data):
        try:
            data = str(data).rstrip('(0123456789)')
            self.outcols[self.OutTitle['學生科系代號']] = Major[data]
        except:
            self.outcols[self.OutTitle['學生科系代號']] = -998
    def convert(self, field_name, data):
        if field_name == '生日':
            return self.birth(data)
        #elif field_name == "學號":
        #    return self.en_id(data)
        elif field_name == "性別":
            return self.gender(data)
        elif field_name == "戶籍地址":
            return self.county(data)
        elif field_name == "年級":
            self.outcols[self.OutTitle['學生年級']] = data
            return 
        elif field_name == "國籍":
            return self.nationality(data)
        elif field_name == "狀態":
            return self.status(data)
        elif field_name == "減免類別":
            return self.identity(data)
        elif field_name == "入學身份別":
            return self.identity(data)
        elif field_name == '入學年月':
            return self.entry_date(data)
        elif field_name == '部別':
            return self.department(data)
        elif field_name == '系所':
            return self.major(data)
        
    # Build a dict to map a city name to its code
    # (This can be placed as a global variable so that you
    # don't need to create the table each time.)
def Read_Map(path):
    with open(path + 'county.csv', 'r') as csv:
        for info in csv.readlines():
            info = info[:-1].split(',')
            County[info[0]] = info[1]
    with open(path + 'major.csv', 'r') as csv:
        for info in csv.readlines():
            info = info[:-1].split(',')
            Major[info[0]] = info[1]
    with open(path +'identity.csv', 'r') as csv:
        for info in csv.readlines():
            info = info[:-1].split(',')
            Identity[info[1]] = info[0]

def main(entry_day ,infile, path = ''):
    Read_Map(path)
    T = Worker( entry_day, infile, 'template.xlsx')
    T.start()
    T.join()
    return entry_day 
    
if __name__ == "__main__":
    main('2024-03-07','static/orig.xlsx')
